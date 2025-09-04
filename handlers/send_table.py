from config import dp, db
from aiogram import types
from aiogram.types import InputFile, InlineKeyboardMarkup, InlineKeyboardButton
import pandas as pd

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


async def send_table(message: types.Message):
    # Создаем книгу Excel
    wb = openpyxl.Workbook()

    # --- Стили ---
    header_font = Font(bold=True, color="FFFFFF")  # Белый жирный текст
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синяя заливка
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # --- Лист1 (метаданные) ---
    sheet1 = wb.active
    sheet1.title = "Лист1"
    sheet1["A1"] = "Дата выгрузки:"
    sheet1["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet1["A1"].font = Font(bold=True)

    # --- Лист "Пользователи и заказы" ---
    sheet2 = wb.create_sheet("Пользователи и заказы")

    # Заголовки
    headers = ["№", "ID", "Дата действия", "Участвуют в бонусной программе?", "ФИО", "Телефон",
               "Сумма показанных баллов", "Анкета состояния", "", "", "Анкета вкусовых предпочтений",
               "", "", "", "", "", "Заказ", "", "Отзыв", "", "Теги (на будущее, чтобы делать рассылку в тг)"]
    sheet2.append(headers)

    # Подзаголовки
    sub_headers = ["", "", "", "", "", "", "", "Настроение", "Голод", "Стиль", "Пол", "Возраст",
                   "Стиль питания", "Ккал", "Не любит", "Любит", "Название позиций", "Сумма чека", "Есть/Нет", "", ""]
    sheet2.append(sub_headers)

    data_rows = []
    # Получаем все записи из таблицы logging
    logging_data = db.get_logging_all()
    i = 1
    
    for log_entry in logging_data:
        # log_entry структура: (id, user_id, date, fio, phone, mood, hungry, style, sex, age, dish_style, ccal, dislike, like, order, sum)
        row = []
        row.append(i)  # №
        row.append(log_entry[1])  # ID пользователя
        row.append(log_entry[2])  # Дата действия
        row.append('Нет')  # Участвуют в бонусной программе?
        row.append(log_entry[3] if log_entry[3] else '-')  # ФИО
        row.append(log_entry[4] if log_entry[4] else '-')  # Телефон
        row.append(0)  # Сумма показанных баллов
        row.append(log_entry[5] if log_entry[5] else '-')  # Настроение
        row.append(log_entry[6] if log_entry[6] else '-')  # Голод
        row.append(log_entry[7] if log_entry[7] else '-')  # Стиль
        row.append(log_entry[8] if log_entry[8] else '-')  # Пол
        row.append(log_entry[9] if log_entry[9] else '-')  # Возраст
        row.append(log_entry[10] if log_entry[10] else '-')  # Стиль питания
        row.append(log_entry[11] if log_entry[11] else '-')  # Ккал
        row.append(log_entry[12] if log_entry[12] else '-')  # Не любит
        row.append(log_entry[13] if log_entry[13] else '-')  # Любит
        row.append(log_entry[14] if log_entry[14] else '-')  # Название позиций (заказ)
        row.append(log_entry[15] if log_entry[15] else '-')  # Сумма чека
        row.append('Нет')  # Есть/Нет (отзыв)
        row.append('')  # Пустое поле
        row.append('')  # Теги (на будущее)
        
        data_rows.append(row)
        i += 1

    for row in data_rows:
        sheet2.append(row)

    # --- Лист "Отзывы" ---
    sheet3 = wb.create_sheet("Отзывы")

    # Заголовки
    sheet3.append(["Дата комментария", "Автор", "Отзыв f2m", "Отзыв на блюдо", "", ""])
    sheet3.append(["", "ID телеграм", "", "Наименование", "Оценка", "Комментарий"])

    # Данные
    reviews = [
        ["2025-07-14 16:05:02", 1456241115, "Удобно пользоваться", "-", "-", "-"],
        ["2025-07-14 16:12:15", 1004320969, "Большой выбор блюд", "Суп с морепродуктами", 3, "-"],
        ["2025-07-14 16:14:15", 1004320969, "Большой выбор блюд", "Каша овсяная", 4, "Неплохо"]
    ]
    for review in reviews:
        sheet3.append(review)

    # --- Применяем стили ко всем листам ---
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Ширина колонок (подбирается индивидуально)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Заливка и шрифт заголовков
        for row in ws.iter_rows(max_row=2):  # Первые две строки - заголовки
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Границы для данных
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Сохраняем файл
    filename = f"full_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"Файл сохранен: {filename}")
    keyboard = InlineKeyboardMarkup()
    keyboard.add(InlineKeyboardButton("⬅️ Назад", callback_data="back_to_boss_menu"))
    await message.answer_document(
        InputFile(filename),
        caption="Вот ваш отчёт 📊",
        reply_markup=keyboard
    )



    # Данные теперь берутся из таблицы logging
    # Все необходимые данные уже включены в основной лист "Пользователи и заказы"
